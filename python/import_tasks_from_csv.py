# -*- coding: utf-8 -*-

# python 2.7 version

import json
import requests
import urlparse
from requests.auth import HTTPBasicAuth
import conf_dahmakan as apiconf
import baserideapiv2 as api2
import csv
import datetime
from openpyxl import load_workbook
import urllib

pattern = {
  "A" : "Email",
  "B" : "Type",
  "C" : "Comment",
  "D" : "Issue Log",
  "E" : "Rider",
  "F" : "Delivery Time",
  "G" : "Item",
  "H" : "Add-Ons",
  "I" : "Delivery Date",
  "J" : "Type",
  "K" : "Quantity per Item",
  "L" : "Price per Item",
  "M" : "Total per Item",
  "N" : "First Name",
  "O" : "Last Name",
  "P" : "Phone",
  "Q" : "Address",
  "R" : "LatLong",
  "S" : "DiscountCode",
  "T" : "OrderNumber",
  "U" : "OrderTimeStamp",
  "V" : "Order Status",
  "W" : "Financial Status",
  "X" : "Order Source"
}

#pattern from file v2
pattern = {
  "A" : "Order No",
  "B" : "Location",
  "C" : "Reach",
  "D" : "Leave",
  "E" : "Expected delivery",
  "F" : "Boxes",
  "G" : "Add-Ons",
  "H" : "Customer Name",
  "I" : "Phone",
  "J" : "To Collect",
  "K" : "Note",
  "L" : "Rider",
  "M" : "LatLong",
  "N" : "Delivery Date" 
}

matching_rules = {
  "name" : "Delivery",
  "description" : "G",
  "address" : "Q"
}

#try:
if True:
    rp_api = api2.BaserideRoutePointApi(apiconf.domain, apiconf.redirect_uri, apiconf.login, apiconf.password, apiconf.client_id, apiconf.client_secret)
    t_api = api2.BaserideTaskApi(apiconf.domain, apiconf.redirect_uri, apiconf.login, apiconf.password, apiconf.client_id, apiconf.client_secret)
    up_api = api2.BaserideUserProfileApi(apiconf.domain, apiconf.redirect_uri, apiconf.login, apiconf.password, apiconf.client_id, apiconf.client_secret)
    rp_api.get_token()
    t_api.set_token(rp_api.access_token)
    up_api.set_token(rp_api.access_token)
    wb = load_workbook(filename = 'CloudTrack import v2 (2).xlsx')
    ws = wb['Sheet1']
    i = 0
    for row in ws.rows:
        i += 1
        if i > 1:
            to_collect = ws["J"+str(i)].value
            boxes = ws["F"+str(i)].value
            addons = ws["G"+str(i)].value
            note = ws["K"+str(i)].value
            reach = ws["C"+str(i)].value
            leave = ws["D"+str(i)].value
            order_number = ws["A"+str(i)].value

            # delivery_datetime = ws["F"+str(i)].value
            #delivery_date = ws["N"+str(i)].value.date() #delivery_datetime.date()
            #delivery_time = ws["C"+str(i)].value.time() #delivery_datetime.time()
            #item = ws["G" + str(i)].value
            #first_name = ws["N" + str(i)].value
            #last_name = ws["O" + str(i)].value
            phone = str(ws["I" + str(i)].value)
            address = ws["B" + str(i)].value
            latlong = ws["M" + str(i)].value
            lat = latlong.split(',')[0]
            lon = latlong.split(',')[1]
            rider = ws["L"+str(i)].value
            #order_number = ws["T"+str(i)].value
            outer_id = apiconf.outer_id_template.format(order_number)
            print "Checking task {}".format(outer_id)
            task_exists_obj = t_api.check_if_exist({"outer_id": outer_id})
            #full_name = first_name + " " + last_name
            full_name = ws["H" + str(i)].value
            if len(task_exists_obj) > 0 :
                print "Task already exists in the system {}".format(outer_id)
            else:
                params = {
                  "lat": lat,
                  "lon": lon,
                  "address" : urllib.quote_plus(address),
                  "phone" : urllib.quote_plus(phone),
                  "name" : urllib.quote_plus(full_name) #first_name + " " + last_name)
                }
                objs = rp_api.check_if_exist(params)
                user_profile_params = {
                  "username" : urllib.quote_plus(str(rider))
                }
                rider_exist = False
                if rider is None:
                    print "Empty rider cell. Check import file. Skipping line {}".format(i)
                else: 
                    up = up_api.check_if_exist(user_profile_params)
                    if len(up) == 0:
                        print "Rider with username {} doesn't exist in the system. First, create such user in admin interface of Baseride".format(rider)
                    else:
                        print "New Job for rider {}".format(up[0]["username"])
                        rider_exist = True
                        rider_uri = up[0]["resource_uri"]

                if rider_exist:
                    if len(objs) == 0: #new route point
                        print "point with such parameters not found. creating new point..."
                        params = {
                          "lat": lat,
                          "lon": lon,
                          "address" : address,
                          "phone" : phone,
                          "name" : full_name #first_name + " " + last_name
                        }
                        rp = params
                        rp["enterprise"] = apiconf.enterprise_url
                        d = rp_api.create_object(rp)
                        print d["id"]
                        route_point_uri = d["resource_uri"]
                    else:
                        route_point_uri = objs[0]["resource_uri"]
                        print "found existing point with id:{}, attaching task to it...".format(route_point_uri)
                    if to_collect is None:
                        to_collect = ""
                    if boxes is None:
                        boxes = ""
                    if addons is None:
                        addons = ""
                    if full_name is None:
                        full_name = ""
                    if phone is None:
                        phone = ""
                    if address is None:
                        address = ""
                    description = "To collect: " + to_collect
                    description += "\nBoxes: " + str(boxes)
                    description += "\nAddons: " + addons
                    description += "\nOrder No: " + str(order_number)
                    description += "\nClient" + full_name
                    description += "\nPhone: " + str(phone)
                    description += "\nAddress: " + address

                    task_params = {
                       "name" : "#"+str(order_number) + " " + address,
                       "outer_id" : outer_id,
                       "target" : route_point_uri,
                       "enterprise" : apiconf.enterprise_url,
                       "performer" : rider_uri,
                       "description" : description
                    }

                    t = t_api.create_object(task_params)
                    print "task {} created".format(t["id"])

#except BaseException, ex:
#    print 'ERR '+str(ex)
